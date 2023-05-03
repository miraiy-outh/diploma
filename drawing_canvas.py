from connection import *
from generating import *
from collections import deque
from tkinter import *
from tkinter import messagebox
from tkinter import ttk
from tkinter import filedialog
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import matplotlib.patches as ptch
from matplotlib import animation
import numpy as np
import csv
import openpyxl

# добавление 1 фрейма с вводом номера партии
def pack_frame1():
    global frame1
    global entry_p
    frame1 = Frame(window)
    frame1.pack()

    label_p = Label(frame1, text="Введите номер партии:", font=('Inter', 20))
    entry_p = Entry(frame1, width=3, font=('Inter', 20), bd=1, highlightthickness=1)
    entry_p.config(highlightbackground = 'black', highlightcolor= 'black')
    button_p = Button(frame1, text='Далее', font=('Inter', 20), bg='lightgrey', bd=1, command=get_entry_p_value)
    label_p.grid(row=1, column=1, pady=50, padx=10)
    entry_p.grid(row=1, column=2)
    button_p.grid(row=2, column=1, sticky=W, padx=10)

# добавление 2 фрейма с вводом параметров новой партии
def pack_frame2():
    global frame2
    global entry_area_s_min
    global entry_area_s_max
    global entry_area_min
    global entry_area_max
    global entry_area_blister_min
    global entry_area_blister_max
    global entry_num

    frame2 = Frame(window)
    frame2.pack()
    label_area_s = Label(frame2, text="Введите диапазон площади в среднем:", font=('Inter', 20))
    entry_area_s_min = Entry(frame2, width=8, font=('Inter', 20), bd=1, highlightthickness=1)
    entry_area_s_min.config(highlightbackground = 'black', highlightcolor= 'black')
    entry_area_s_max = Entry(frame2, width=8, font=('Inter', 20), bd=1, highlightthickness=1)
    entry_area_s_max.config(highlightbackground = 'black', highlightcolor= 'black')

    label_area = Label(frame2, text="Введите диапазон площади в общем:", font=('Inter', 20))
    entry_area_min = Entry(frame2, width=8, font=('Inter', 20), bd=1, highlightthickness=1)
    entry_area_min.config(highlightbackground = 'black', highlightcolor= 'black')
    entry_area_max = Entry(frame2, width=8, font=('Inter', 20), bd=1, highlightthickness=1)
    entry_area_max.config(highlightbackground = 'black', highlightcolor= 'black')

    label_area_blister = Label(frame2, text="Введите диапазон площади блистера:", font=('Inter', 20))
    entry_area_blister_min = Entry(frame2, width=8, font=('Inter', 20), bd=1, highlightthickness=1)
    entry_area_blister_min.config(highlightbackground = 'black', highlightcolor= 'black')
    entry_area_blister_max = Entry(frame2, width=8, font=('Inter', 20), bd=1, highlightthickness=1)
    entry_area_blister_max.config(highlightbackground = 'black', highlightcolor= 'black')

    label_num = Label(frame2, text="Введите количество областей:", font=('Inter', 20))
    entry_num = Entry(frame2, width=8, font=('Inter', 20), bd=1, highlightthickness=1)
    entry_num.config(highlightbackground = 'black', highlightcolor= 'black')

    button_p2 = Button(frame2, text='Добавить партию', font=('Inter', 20), bg='lightgrey', bd=1, command=get_entries_values)
    button_p2_1 = Button(frame2, text='Назад', font=('Inter', 20), bg='lightgrey', bd=1, command=button_frame2)

    label_area_s.grid(row=1, column=1, pady=10, sticky=W)
    entry_area_s_min.grid(row=1, column=2, padx=10, pady=10)
    entry_area_s_max.grid(row=1, column=3, pady=10)

    label_area.grid(row=2, column=1, pady=10, sticky=W)
    entry_area_min.grid(row=2, column=2, padx=10, pady=10)
    entry_area_max.grid(row=2, column=3, pady=10)

    label_area_blister.grid(row=3, column=1, pady=10, sticky=W)
    entry_area_blister_min.grid(row=3, column=2, padx=10, pady=10)
    entry_area_blister_max.grid(row=3, column=3, pady=10)

    label_num.grid(row=4, column=1, pady=10, sticky=W)
    entry_num.grid(row=4, column=2, padx=10, pady=10)

    button_p2.grid(row=5, column=1, pady=10, sticky=W)
    button_p2_1.grid(row=5, column=3, pady=10, sticky=E)

# добавление фрейма со статистикой
def pack_frame3():
    global frame3
    global tab_control
    global canvas
    global canvas1
    global canvas2
    global info_frame
    global tree
    global text
    global anim
    global anim1
    global fig
    global fig1
    global ax1
    global ax2
    global ax3
    global ax4
    global ax5
    global ax6
    global strs # количество объектов на графиках
    global x #очередь
    strs = 20 # количество записей на графике
    
    #заполнение очереди последними strs объектами
    tmp20 = f'SELECT COUNT(ID_object) FROM Object WHERE ID_jugement={type_of_pills};'
    tmp20 = execute_read_query(connection, tmp20)[0][0]
    x = deque(list(range(tmp20 - strs, tmp20)), maxlen=strs)

    frame3 = Frame(window)
    frame3.pack(fill='both', expand=True)

    # добавление меню вкладок
    style = ttk.Style()
    style.configure('TFrame', background='white')
    tab_control = ttk.Notebook(frame3)  
    tab1 = ttk.Frame(tab_control, style='TFrame')
    tab2 = ttk.Frame(tab_control, style='TFrame')
    tab3 = ttk.Frame(tab_control, style='TFrame')
    tab_control.add(tab1, text='Статистика параметров') 
    tab_control.add(tab2, text='Статистика брака') 
    tab_control.add(tab3, text='Подробнее о браке') 
    tab_control.pack(expand=1, fill='both')
    button_export_frame = Frame(tab1, bg='white')
    button_export_frame.pack(fill='both', side=TOP, pady=10)
    button_export = Button(master=button_export_frame, text='Сохранить в файл', font=('Inter', 10), bg='lightgrey', bd=1, command=button_export_click)
    button_export.pack(anchor=NE, padx=20, side=RIGHT)
    button = Button(master=button_export_frame, text='Сменить партию', font=('Inter', 10), bg='lightgrey', bd=1, command=button_click)
    button.pack(anchor=NE, padx=20, side=RIGHT)
    button1 = Button(master=tab2, text='Сменить партию', font=('Inter', 10), bg='lightgrey', bd=1, command=button_click)
    button1.pack(anchor=NE, padx=20, pady=10)

    # создание полей для графиков
    fig, (ax1, ax2, ax3, ax4) = plt.subplots(nrows=4, ncols=1, figsize=(8, 30))
    plt.subplots_adjust(hspace=0.5)
    anim = animation.FuncAnimation(fig, adding, interval = 1000)
    canvas = FigureCanvasTkAgg(fig, master=tab1)
    canvas.get_tk_widget().pack(fill='both', expand=True)

    # создание полей для графика брака
    fig1, (ax5, ax6) = plt.subplots(2, 1, figsize=(10, 10))
    plt.subplots_adjust(hspace=0.5)
    anim1 = animation.FuncAnimation(fig1, adding1, interval = 1000)
    canvas1 = FigureCanvasTkAgg(fig1, master=tab2)
    canvas1.get_tk_widget().pack(fill='both', expand=True)

    # создание 3 вкладки
    canvas2 = Canvas(tab3, bg='white', bd=0, highlightthickness=0)
    canvas2.pack(fill='both', expand=True)
    # фрейм с кнопками
    button_frame = Frame(canvas2, bg='white')
    button_frame.pack(side=TOP, pady=10)
    button_blister = Button(master=button_frame, text='Форма блистера', font=('Inter', 10), bg='lightgrey', bd=1, command=button_blister_click)
    button_blister.pack(padx=10, side=LEFT)
    button_color = Button(master=button_frame, text='Цвет таблетки', font=('Inter', 10), bg='lightgrey', bd=1, command=button_color_click)
    button_color.pack(padx=10, side=LEFT)
    button_scratch = Button(master=button_frame, text='Скол таблетки', font=('Inter', 10), bg='lightgrey', bd=1, command=button_scratch_click)
    button_scratch.pack(padx=10, side=LEFT)
    button_absence = Button(master=button_frame, text='Отсутствие таблетки', font=('Inter', 10), bg='lightgrey', bd=1, command=button_absence_click)
    button_absence.pack(padx=10, side=LEFT)
    button_form = Button(master=button_frame, text='Форма таблетки', font=('Inter', 10), bg='lightgrey', bd=1, command=button_form_click)
    button_form.pack(padx=10, side=LEFT)
    button_all = Button(master=button_frame, text='Все', font=('Inter', 10), bg='lightgrey', bd=1, command=button_all_click)
    button_all.pack(padx=10, side=LEFT)
    # фреймы с информацией
    stat_frame = Frame(canvas2, bg='white')
    stat_frame.pack(side=TOP, pady=10)
    text = Text(master=stat_frame, font=('Inter', 20), height=1, bd=0, width=20)
    text.pack(fill=BOTH, expand=True)
    info_frame = Frame(canvas2, bg='white')
    info_frame.pack(side=TOP, pady=10)
    scrollbar = Scrollbar(info_frame)
    scrollbar.pack(side=RIGHT, fill=Y)
    columns = ('number_of_pill', 'area')
    tree = ttk.Treeview(master=info_frame, columns=columns, show='headings',  yscrollcommand=scrollbar.set, height=20)
    tree.pack(fill=BOTH, side=LEFT)
    tree.heading('number_of_pill', text='Номер таблетки', anchor=N)
    tree.heading('area', text='Площадь таблетки', anchor=N)
    tree.column("#1", anchor=N, stretch=True)
    tree.column("#2", anchor=N, stretch=True)
    scrollbar.config(command=tree.yview)
    # фрейм с кнопкой экспорта
    export_frame = Frame(canvas2, bg='white')
    export_frame.pack(side=TOP, pady=10)
    button_export_error = Button(master=export_frame, text='Сохранить в файл', font=('Inter', 10), bg='lightgrey', bd=1, command=button_export_error_click)
    button_export_error.pack(padx=10, side=RIGHT)

    tab_control.bind('<<NotebookTabChanged>>', on_tab_switch)

# функции отработки нажатия кнопок
def button_frame2():
    frame2.destroy()
    pack_frame1()

def button_click():
    frame3.destroy()
    pack_frame1()

def button_export_click():
    tmp = f'SELECT ID_object, jugement, area, total_area, area_blister, number_of_labels, time FROM Object WHERE ID_jugement={type_of_pills};'
    tmp = execute_read_query(connection, tmp)
    folder_selected = filedialog.askdirectory()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    head_mas = ['ID объекта', 'Результат проверки', 'Площадь в среднем', 'Площадь в общем', 'Площадь блистера', 'Количество таблеток', 'Время']
    for i in range(len(head_mas)):
        worksheet.cell(row=1, column=i + 1, value=head_mas[i])
    for row_index, row_data in enumerate(tmp, start=2):
        for col_index, cell_data in enumerate(row_data, start=1):
            worksheet.cell(row=row_index, column=col_index, value=cell_data)
    workbook.save(f'{folder_selected}/object_info.xlsx')

def button_export_error_click():
    tmp = f'SELECT ID_error_pill, number_of_pill, area_pill, ID_error FROM Error_pill WHERE ID_jugement={type_of_pills};'
    tmp = execute_read_query(connection, tmp)
    folder_selected = filedialog.askdirectory()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    head_mas = ['ID записи', 'Номер таблетки', 'Площадь таблетки', 'ID ошибки']
    for i in range(len(head_mas)):
        worksheet.cell(row=1, column=i + 1, value=head_mas[i])
    for row_index, row_data in enumerate(tmp, start=2):
        for col_index, cell_data in enumerate(row_data, start=1):
            worksheet.cell(row=row_index, column=col_index, value=cell_data)
    workbook.save(f'{folder_selected}/error_info.xlsx')

def button_blister_click():
    text.delete('1.0', END)
    tree.delete(*tree.get_children())
    text_info('1')

def button_color_click():
    text.delete('1.0', END)
    tree.delete(*tree.get_children())
    text_info('2')

def button_scratch_click():
    text.delete('1.0', END)
    tree.delete(*tree.get_children())
    text_info('3')

def button_absence_click():
    text.delete('1.0', END)
    tree.delete(*tree.get_children())
    text_info('5')

def button_form_click():
    text.delete('1.0', END)
    tree.delete(*tree.get_children())
    text_info('4')

def button_all_click():
    text.delete('1.0', END)
    tree.delete(*tree.get_children())
    text_info('all')

# вывод информации в текстовое поле
def text_info(param):
    if param == 'all':
        tmp = f'SELECT number_of_pill, area_pill FROM Error_pill WHERE ID_jugement={type_of_pills};'
        tmp = execute_read_query(connection, tmp)
        tmp1 = f'SELECT number_of_pill FROM error_pill GROUP BY number_of_pill ORDER BY COUNT(number_of_pill) DESC LIMIT 1;'
        tmp1 = execute_read_query(connection, tmp1)[0][0]
    else:
        tmp = f'SELECT number_of_pill, area_pill FROM Error_pill WHERE ID_error={param} AND ID_jugement={type_of_pills};'
        tmp = execute_read_query(connection, tmp)
        tmp1 = f'SELECT number_of_pill FROM error_pill WHERE ID_error={param} AND ID_jugement={type_of_pills} GROUP BY number_of_pill ORDER BY COUNT(number_of_pill) DESC LIMIT 1;'
        tmp1 = execute_read_query(connection, tmp1)[0][0]
    text.insert('1.0', f'Больше всего ошибок: {tmp1}')
    for row in tmp:
        tree.insert('', END, values=row)
        

# получение значений партии и вывод нового фрейма
def get_entry_p_value():
    global type_of_pills # номер партии

    if entry_p.get().isdigit():
        type_of_pills = int(entry_p.get())
        tmp = f'SELECT * FROM Jugement WHERE ID_jugement={type_of_pills};'
        tmp = execute_read_query(connection, tmp)
        if tmp != []:
            frame1.destroy()
            pack_frame3()
        else:
            frame1.destroy()
            pack_frame2()
    else:
        messagebox.showinfo('Ошибка', 'Введите корректный номер партии!')

# получение значений новой партии и запись в БД
def get_entries_values():
    if entry_area_s_min.get().isdigit() and entry_area_s_max.get().isdigit() and entry_area_min.get().isdigit() and entry_area_max.get().isdigit() and entry_area_blister_min.get().isdigit() and entry_area_blister_max.get().isdigit() and entry_num.get().isdigit():
        tamin = int(entry_area_min.get())
        tamax = int(entry_area_max.get())
        amin = int(entry_area_s_min.get())
        amax = int(entry_area_s_max.get())
        nums = int(entry_num.get())
        bmin = int(entry_area_blister_min.get())
        bmax = int(entry_area_blister_max.get())
        tmp = f'INSERT INTO jugement (ID_jugement, area_min, area_max, total_area_min, total_area_max, number_of_labels_min, number_of_labels_max, area_blister_min, area_blister_max) VALUES ({type_of_pills},{amin}, {amax}, {tamin}, {tamax}, {nums}, {nums}, {bmin}, {bmax});'
        tmp = execute_read_query(connection, tmp)
        gnrt(tamin, tamax, amin, amax, nums, nums, bmin, bmax, type_of_pills)
        pills(tamin, tamax, amin, amax, nums, nums, bmin, bmax, type_of_pills)
        frame2.destroy()
        pack_frame3()
        
    else:
        messagebox.showinfo('Ошибка', 'Введите корректные параметры партии!')

# получение допустимых параметров партии
def get_jugement():
    tmp = f'SELECT * FROM Jugement WHERE ID_jugement={type_of_pills};'
    tmp = execute_read_query(connection, tmp)
    areamin = tmp[0][1]
    areamax = tmp[0][2]
    totalareamin = tmp[0][3]
    totalareamax = tmp[0][4]
    numbersmin = tmp[0][5]
    numbersmax = tmp[0][6]
    blistermin = tmp[0][7]
    blistermax = tmp[0][8]
    id_jugement = tmp[0][0]
    pills(totalareamin, totalareamax, areamin, areamax, numbersmin, numbersmax, blistermin, blistermax, id_jugement)
    return totalareamin, totalareamax, areamin, areamax, numbersmin, numbersmax, blistermin, blistermax

# функция для получения последних strs записей из бд типа type_of_pills
def massives(strs, type_of_pills):
    tmp = f'SELECT * FROM (SELECT * FROM Object WHERE ID_jugement={type_of_pills} ORDER BY ID_object DESC LIMIT {strs}) t ORDER BY ID_object;'
    tmp = execute_read_query(connection, tmp)

    numbers_mas = []
    area_mas = []
    totalarea_mas = []
    areapills_mass = []

    for i in range(strs):
        numbers_mas.append(tmp[i][5])
        area_mas.append(tmp[i][2])
        totalarea_mas.append(tmp[i][3])
        areapills_mass.append(tmp[i][4])
    
    tmp0 = f'SELECT COUNT(ID_object) FROM Object WHERE jugement=0 AND ID_jugement={type_of_pills};'
    tmp0 = execute_read_query(connection, tmp0)[0][0]
    tmp1 = f'SELECT COUNT(ID_object) FROM Object WHERE jugement=-1 AND ID_jugement={type_of_pills};'
    tmp1 = execute_read_query(connection, tmp1)[0][0]
    pers = (tmp1 / (tmp0 + tmp1)) * 100
    pers = format(pers, '.2f')
    cnt = tmp0 + tmp1

    id_obj = f'SELECT COUNT(ID_object) FROM Object WHERE ID_jugement={type_of_pills};'
    id_obj = execute_read_query(connection, id_obj)[0][0]

    last_hour= f'SELECT COUNT(ID_object) FROM Object WHERE time > DATE_ADD(NOW(), INTERVAL -1 HOUR) AND ID_jugement={type_of_pills};'
    last_hour = execute_read_query(connection, last_hour)[0][0]
    return numbers_mas, area_mas, totalarea_mas, areapills_mass, pers, cnt, id_obj, last_hour

# функция для отрисовки графиков
def draw(ax, title, x, y, x1, y1, y2):
    ax.clear()
    draw_brdrs(ax, x1 - strs - 1, y1, y2)
    ax.plot(x, y)
    ax.relim()
    ax.autoscale_view(True, True, True)
    ax.set_title(f'{title}')
    ax.xaxis.set_major_locator(ticker.MultipleLocator(1))

# функция для отрисовки допустимых границ
def draw_brdrs(ax, x1, y1, y2):
    ax.add_patch(
     ptch.Rectangle(
        (x1, y1),
        strs,
        y2-y1,
        facecolor = 'red',
        fill=True,
        alpha=0.2
     ) )

# функция для обновления графиков статистики
def adding(i):
    totalareamin, totalareamax, areamin, areamax, numbersmin, numbersmax, blistermin, blistermax = get_jugement()
    numbers_mas, area_mas, totalarea_mas, areapills_mass, pers, cnt, id_obj, hours = massives(strs, type_of_pills)

    fig.suptitle(f'Тип партии: {type_of_pills}. Всего изделий: {cnt}, процент брака: {pers}%. Произведено за последний час: {hours}')

    draw(ax1, 'Количество объектов', x, numbers_mas, id_obj, numbersmin, numbersmax)
    draw(ax2, 'Площадь области в среднем', x, area_mas, id_obj, areamin, areamax)
    draw(ax3, 'Площадь всех областей', x, totalarea_mas, id_obj, totalareamin, totalareamax)
    draw(ax4, 'Площадь блистера', x, areapills_mass, id_obj, blistermin, blistermax)

    x.append(x[strs - 1] + 1)

# функция для получения количества брака по типу
def get_error():
    tmp_mas = []
    for i in range(5):
        tmp1 = f'SELECT COUNT(ID_error_pill) FROM Error_pill WHERE ID_error={i + 1} AND ID_jugement={type_of_pills};'
        tmp_mas.append(execute_read_query(connection, tmp1)[0][0])

    return tmp_mas

# функция для получения количества брака по номеру таблетки
def get_number_pill():
    tmp = f'SELECT * FROM Jugement WHERE ID_jugement={type_of_pills};'
    tmp = execute_read_query(connection, tmp)
    numbers = tmp[0][5]
    tmp_mas = []
    for i in range(numbers):
        tmp1 = f'SELECT COUNT(ID_error_pill) FROM Error_pill WHERE number_of_pill={i + 1} AND ID_jugement={type_of_pills};'
        tmp_mas.append(execute_read_query(connection, tmp1)[0][0])

    return tmp_mas

# функция для обновления графика ошибок
def adding1(i):
    totalareamin, totalareamax, areamin, areamax, numbersmin, numbersmax, blistermin, blistermax = get_jugement()
    numbers_mas, area_mas, totalarea_mas, areapills_mass, pers, cnt, id_obj, hours = massives(strs, type_of_pills)
    fig1.suptitle(f'Тип партии: {type_of_pills}. Всего изделий: {cnt}, процент брака: {pers}%. Произведено за последний час: {hours}')
    mas = get_error()
    ax5.clear()
    a = np.array(['Форма блистера', 'Цвет таблетки', 'Скол таблетки', 'Отсутствие таблетки', 'Форма таблетки'])
    b = np.array([mas[0], mas[1], mas[2], mas[3], mas[4]])
    ax5.bar(a, b)
    ax5.set_title('Количество брака по типу')

    mas = get_number_pill()
    ax6.clear()
    c = np.array([1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
    d = np.array([mas[0], mas[1], mas[2], mas[3], mas[4], mas[5], mas[6], mas[7], mas[8], mas[9]])
    ax6.bar(c, d)
    ax6.xaxis.set_major_locator(ticker.MultipleLocator(1))
    ax6.yaxis.set_major_locator(ticker.MultipleLocator(1))
    ax6.set_title('Количество брака по номеру таблетки')

    x.append(x[strs - 1] + 1)

# функция смены анимаций на разных вкладках
def on_tab_switch(event):
    if tab_control.index('current') == 0:
        canvas1.unbind('<Button-1>')
        canvas1.unbind('<Key>')
        canvas2.unbind('<Button-1>')
        canvas2.unbind('<Key>')
    elif tab_control.index('current') == 1:
        canvas.unbind('<Button-1>')
        canvas.unbind('<Key>')
        canvas2.unbind('<Button-1>')
        canvas2.unbind('<Key>')
    else:
        canvas.unbind('<Button-1>')
        canvas.unbind('<Key>')
        canvas1.unbind('<Button-1>')
        canvas1.unbind('<Key>')

# создание окна
window = Tk()
window.title("Статистика")
window.geometry('1200x700')
pack_frame1()
window.mainloop()